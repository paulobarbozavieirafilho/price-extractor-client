from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


def load_pending_reviews(editor_path: Path, limit: int = 300) -> list[dict]:
    if not editor_path.exists():
        return []
    try:
        df = pd.read_excel(editor_path, sheet_name="Review")
    except Exception:
        return []

    if df.empty:
        return []

    if "status" not in df.columns:
        df["status"] = ""
    if "fingerprint" not in df.columns:
        return []

    status = df["status"].fillna("").astype(str).str.strip().str.upper()
    pending = df[status.isin({"", "PENDING", "REVIEW"})].copy()
    if pending.empty:
        return []

    cols = [
        "fingerprint",
        "xProd",
        "xProd_norm",
        "uCom",
        "qCom",
        "vUnCom",
        "suggested_sku_id",
        "suggested_sku_name_canonical",
        "suggested_base_measure",
        "confidence",
        "status",
        "chosen_sku_id",
        "notes",
    ]
    for col in cols:
        if col not in pending.columns:
            pending[col] = ""

    result = pending[cols].fillna("").head(limit).to_dict(orient="records")
    return result


def apply_review_decision(
    editor_path: Path,
    *,
    fingerprint: str,
    action: str,
    chosen_sku_id: str | None = None,
    notes: str | None = None,
) -> bool:
    if not editor_path.exists():
        return False

    wb = load_workbook(editor_path)
    if "Review" not in wb.sheetnames:
        return False

    ws = wb["Review"]
    headers: dict[str, int] = {}
    for col_idx in range(1, ws.max_column + 1):
        key = ws.cell(row=1, column=col_idx).value
        if key:
            headers[str(key)] = col_idx

    required = {"fingerprint", "status"}
    if not required.issubset(set(headers.keys())):
        return False

    fp_col = headers["fingerprint"]
    status_col = headers["status"]
    chosen_col = headers.get("chosen_sku_id")
    notes_col = headers.get("notes")
    updated_col = headers.get("updated_at")

    target_row = None
    target_fingerprint = str(fingerprint).strip()
    for row_idx in range(2, ws.max_row + 1):
        value = ws.cell(row=row_idx, column=fp_col).value
        if value is None:
            continue
        if str(value).strip() == target_fingerprint:
            target_row = row_idx
            break

    if target_row is None:
        return False

    action_norm = action.strip().lower()
    if action_norm == "approve":
        ws.cell(row=target_row, column=status_col, value="APPROVED")
        if chosen_col:
            final_sku = (chosen_sku_id or "").strip()
            if final_sku:
                ws.cell(row=target_row, column=chosen_col, value=final_sku)
    elif action_norm == "ignore":
        ws.cell(row=target_row, column=status_col, value="IGNORE")
    else:
        return False

    if notes_col and notes is not None:
        ws.cell(row=target_row, column=notes_col, value=notes.strip())
    if updated_col:
        ws.cell(
            row=target_row,
            column=updated_col,
            value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        )

    wb.save(editor_path)
    return True

